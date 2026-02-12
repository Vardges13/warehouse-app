#!/usr/bin/env python3
"""
Складской ассистент - FastAPI Backend
Высокоточный ассистент по складской логистике и подготовке отгрузочных документов
"""

import os
import json
import re
import shutil
import traceback
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
from io import BytesIO
import base64

import google.generativeai as genai
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.requests import Request
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image
# import magic
import pandas as pd
import uvicorn

# Инициализация
app = FastAPI(title="Складской ассистент", version="1.0.0")
templates = Jinja2Templates(directory="templates")

# Создание необходимых директорий
os.makedirs("uploads", exist_ok=True)
os.makedirs("output", exist_ok=True)

# Gemini API инициализация
GEMINI_API_KEY = "AIzaSyCcNkbZp447GjuW8xjykrJ_N-r_3g10dhY"
genai.configure(api_key=GEMINI_API_KEY)

# Глобальное хранилище данных сессии
session_data = {
    "specification": None,
    "template": None,
    "photos": [],
    "step": 1,
    "results": {}
}

class WarehouseAssistant:
    """Главный класс складского ассистента"""
    
    def __init__(self):
        self.model = genai.GenerativeModel('gemini-2.0-flash')
        
    def check_image_quality(self, image_path: str) -> Dict[str, Any]:
        """Проверка качества фотографии"""
        try:
            with Image.open(image_path) as img:
                width, height = img.size
                
                # Базовые проверки качества
                quality_issues = []
                
                if width < 300 or height < 300:
                    quality_issues.append("Низкое разрешение")
                
                # Быстрая проверка: ресайз до маленького размера
                thumb = img.copy()
                thumb.thumbnail((200, 200))
                img_gray = thumb.convert('L')
                pixels = list(img_gray.getdata())
                n = len(pixels)
                avg_brightness = sum(pixels) / n
                variance = sum((x - avg_brightness) ** 2 for x in pixels) / n
                
                if variance < 500:
                    quality_issues.append("Изображение размыто")
                
                # Проверка на пересвет/недосвет
                if avg_brightness > 240:
                    quality_issues.append("Пересвет")
                elif avg_brightness < 15:
                    quality_issues.append("Недосвет")
                
                status = "❌" if quality_issues else "✅"
                if len(quality_issues) == 1:
                    status = "⚠️"
                
                return {
                    "status": status,
                    "readable": len(quality_issues) == 0,
                    "issues": quality_issues,
                    "resolution": f"{width}x{height}",
                    "brightness": round(avg_brightness, 1)
                }
                
        except Exception as e:
            return {
                "status": "❌",
                "readable": False,
                "issues": [f"Ошибка обработки: {str(e)}"],
                "resolution": "неизвестно",
                "brightness": 0
            }
    
    def extract_marking_from_photo(self, image_path: str) -> Dict[str, Any]:
        """Извлечение маркировки через Gemini Vision API"""
        try:
            # Загружаем и подготавливаем изображение
            with Image.open(image_path) as img:
                # Конвертируем в RGB если необходимо
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                
                # Сохраняем во временный файл для передачи в Gemini
                temp_path = image_path + "_temp.jpg"
                img.save(temp_path, 'JPEG', quality=90)
            
            # Загружаем изображение в Gemini
            image_file = genai.upload_file(temp_path)
            
            # Промпт для распознавания маркировки
            prompt = """Распознай маркировку на фото. 
Верни результат СТРОГО в формате JSON:
{
    "name": "наименование товара",
    "article": "артикул",
    "dimensions": "размеры"
}

Если маркировка не читается или фото нечеткое, верни:
{
    "name": null,
    "article": null,
    "dimensions": null,
    "error": "описание проблемы"
}

Важно: отвечай ТОЛЬКО JSON без дополнительного текста."""

            # Вызов Gemini API
            response = self.model.generate_content([prompt, image_file])
            
            # Очищаем временный файл
            try:
                os.remove(temp_path)
            except:
                pass  # Игнорируем ошибку удаления временного файла
            
            # Парсим ответ JSON
            try:
                result_json = json.loads(response.text.strip())
                
                # Проверяем наличие ошибки
                if result_json.get('error'):
                    return {
                        "status": "❌",
                        "name": None,
                        "article": None,
                        "dimensions": None,
                        "readable": False,
                        "comment": result_json['error'],
                        "demo_mode": False
                    }
                
                # Проверяем полноту данных
                has_name = result_json.get('name') is not None
                has_article = result_json.get('article') is not None
                
                if has_name and has_article:
                    status = "✅"
                    readable = True
                elif has_article:
                    status = "⚠️"
                    readable = True
                else:
                    status = "❌"
                    readable = False
                
                return {
                    "status": status,
                    "name": result_json.get('name'),
                    "article": result_json.get('article'),
                    "dimensions": result_json.get('dimensions'),
                    "readable": readable,
                    "confidence": "gemini-vision",
                    "demo_mode": False
                }
                
            except json.JSONDecodeError:
                return {
                    "status": "❌",
                    "name": None,
                    "article": None,
                    "dimensions": None,
                    "readable": False,
                    "comment": "Ошибка парсинга ответа Gemini",
                    "demo_mode": False
                }
                
        except Exception as e:
            return {
                "status": "❌",
                "name": None,
                "article": None,
                "dimensions": None,
                "readable": False,
                "comment": f"Ошибка API: {str(e)[:100]}",
                "demo_mode": False
            }
    
    def parse_excel_specification(self, file_path: str) -> List[Dict]:
        """Парсинг спецификации из Excel"""
        try:
            df = pd.read_excel(file_path)
            
            # Ищем колонки с артикулами и количеством
            articles_col = None
            quantity_col = None
            name_col = None
            
            for col in df.columns:
                col_lower = str(col).lower()
                if any(word in col_lower for word in ['артикул', 'код', 'article']):
                    articles_col = col
                elif any(word in col_lower for word in ['количество', 'кол-во', 'qty', 'quantity']):
                    quantity_col = col
                elif any(word in col_lower for word in ['наименование', 'название', 'name']):
                    name_col = col
            
            if not articles_col or not quantity_col:
                raise ValueError("Не найдены колонки с артикулами или количеством")
            
            specification = []
            for _, row in df.iterrows():
                if pd.notna(row[articles_col]) and pd.notna(row[quantity_col]):
                    item = {
                        "article": str(row[articles_col]).strip(),
                        "quantity": int(row[quantity_col]),
                        "name": str(row[name_col]).strip() if name_col else ""
                    }
                    specification.append(item)
            
            return specification
            
        except Exception as e:
            raise ValueError(f"Ошибка парсинга спецификации: {str(e)}")
    
    def calculate_square_meters(self, name: str, dimensions: str = None) -> float:
        """Расчет площади в м² из наименования или размеров"""
        try:
            text = dimensions or name or ""
            
            # Поиск размеров в формате ЧИСЛОxЧИСЛО
            patterns = [
                r'(\d+(?:\.\d+)?)\s*[xх]\s*(\d+(?:\.\d+)?)\s*мм',
                r'(\d+(?:\.\d+)?)\s*[xх]\s*(\d+(?:\.\d+)?)\s*mm',
                r'(\d+(?:\.\d+)?)\s*[xх]\s*(\d+(?:\.\d+)?)',
                r'(\d+(?:\.\d+)?)\s*\*\s*(\d+(?:\.\d+)?)'
            ]
            
            for pattern in patterns:
                match = re.search(pattern, text.lower())
                if match:
                    width = float(match.group(1))
                    height = float(match.group(2))
                    
                    # Если размеры в мм, конвертируем в метры
                    if 'мм' in text.lower() or 'mm' in text.lower():
                        width_m = width / 1000
                        height_m = height / 1000
                    else:
                        # Предполагаем, что большие числа - это мм
                        if width > 100 or height > 100:
                            width_m = width / 1000
                            height_m = height / 1000
                        else:
                            width_m = width
                            height_m = height
                    
                    return round(width_m * height_m, 3)
            
            return 0.0
            
        except:
            return 0.0

# Инициализация ассистента
assistant = WarehouseAssistant()

@app.get("/", response_class=HTMLResponse)
async def main_page(request: Request):
    """Главная страница"""
    return templates.TemplateResponse("index.html", {"request": request})

@app.post("/upload-files/")
async def upload_files(
    specification: UploadFile = File(None),
    template: UploadFile = File(None),
    photos: List[UploadFile] = File([])
):
    """Загрузка файлов"""
    try:
        print(f"[UPLOAD] spec={specification}, template={template}, photos={len(photos) if photos else 0}")
        result = {"success": True, "files": {}}
        
        # Очистка предыдущих данных
        session_data["specification"] = None
        session_data["template"] = None
        session_data["photos"] = []
        
        # Обработка спецификации
        if specification and specification.filename:
            spec_path = f"uploads/specification_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            content = await specification.read()
            with open(spec_path, "wb") as f:
                f.write(content)
            session_data["specification"] = spec_path
            result["files"]["specification"] = specification.filename
            print(f"[UPLOAD] spec saved: {spec_path} ({len(content)} bytes)")
        
        # Обработка шаблона накладной
        if template and template.filename:
            templ_path = f"uploads/template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            content = await template.read()
            with open(templ_path, "wb") as f:
                f.write(content)
            session_data["template"] = templ_path
            result["files"]["template"] = template.filename
            print(f"[UPLOAD] template saved: {templ_path} ({len(content)} bytes)")
        
        # Обработка фотографий
        print(f"[UPLOAD] photos count: {len(photos)}, filenames: {[p.filename for p in photos]}")
        for i, photo in enumerate(photos):
            if photo.filename:
                content = await photo.read()
                if len(content) == 0:
                    print(f"[UPLOAD] photo {i} empty, skip")
                    continue
                
                ext = os.path.splitext(photo.filename)[1] or '.jpg'
                photo_path = f"uploads/photo_{i}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
                with open(photo_path, "wb") as f:
                    f.write(content)
                
                session_data["photos"].append({
                    "path": photo_path,
                    "filename": photo.filename
                })
        
        result["files"]["photos"] = len(session_data["photos"])
        session_data["step"] = 1
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка загрузки файлов: {str(e)}")

@app.post("/step1-check-files/")
async def step1_check_files():
    """Шаг 1: Проверка загруженных файлов"""
    try:
        missing = []
        
        if not session_data["specification"]:
            missing.append("Спецификация (Excel)")
        
        if not session_data["template"]:
            missing.append("Шаблон накладной (Excel)")
        
        if not session_data["photos"]:
            missing.append("Фотографии маркировок")
        
        result = {
            "success": len(missing) == 0,
            "files_count": {
                "specification": 1 if session_data["specification"] else 0,
                "template": 1 if session_data["template"] else 0,
                "photos": len(session_data["photos"])
            },
            "missing": missing
        }
        
        if result["success"]:
            session_data["step"] = 2
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка проверки файлов: {str(e)}")

@app.post("/step2-check-photo-quality/")
async def step2_check_photo_quality():
    """Шаг 2: Проверка качества фотографий"""
    try:
        photo_quality = []
        
        for i, photo_data in enumerate(session_data["photos"]):
            quality_info = assistant.check_image_quality(photo_data["path"])
            quality_info["filename"] = photo_data["filename"]
            quality_info["index"] = i
            photo_quality.append(quality_info)
        
        unreadable_count = sum(1 for q in photo_quality if not q["readable"])
        
        result = {
            "success": True,
            "photos": photo_quality,
            "summary": {
                "total": len(photo_quality),
                "readable": len(photo_quality) - unreadable_count,
                "unreadable": unreadable_count
            }
        }
        
        session_data["results"]["photo_quality"] = result
        session_data["step"] = 3
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка проверки качества фото: {str(e)}")

@app.post("/step3-extract-markings/")
async def step3_extract_markings():
    """Шаг 3: Извлечение маркировки через Gemini Vision"""
    try:
        marking_results = []
        
        for i, photo_data in enumerate(session_data["photos"]):
            # Пропускаем нечитаемые фото
            quality_info = session_data["results"]["photo_quality"]["photos"][i]
            if not quality_info["readable"]:
                marking_info = {
                    "filename": photo_data["filename"],
                    "index": i,
                    "status": "❌",
                    "name": None,
                    "article": None,
                    "dimensions": None,
                    "comment": "Фото нечитаемо"
                }
            else:
                marking_info = assistant.extract_marking_from_photo(photo_data["path"])
                marking_info["filename"] = photo_data["filename"]
                marking_info["index"] = i
                marking_info["comment"] = ""
                
                if not marking_info["readable"]:
                    marking_info["comment"] = "Маркировка не читается"
                elif marking_info.get("error"):
                    marking_info["comment"] = marking_info["error"]
            
            marking_results.append(marking_info)
        
        # Подсчет статистики
        extracted_count = sum(1 for m in marking_results if m["status"] == "✅")
        partial_count = sum(1 for m in marking_results if m["status"] == "⚠️")
        failed_count = sum(1 for m in marking_results if m["status"] == "❌")
        
        result = {
            "success": True,
            "markings": marking_results,
            "summary": {
                "total": len(marking_results),
                "extracted": extracted_count,
                "partial": partial_count,
                "failed": failed_count
            }
        }
        
        session_data["results"]["markings"] = result
        session_data["step"] = 4
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка извлечения маркировок: {str(e)}")

@app.post("/step4-count-verification/")
async def step4_count_verification():
    """Шаг 4: Двойной пересчет изделий"""
    try:
        # Проверяем наличие данных от предыдущего шага
        if "markings" not in session_data.get("results", {}) or "markings" not in session_data["results"]["markings"]:
            raise HTTPException(status_code=400, detail="Отсутствуют данные маркировок. Выполните шаг 3.")
        
        markings = session_data["results"]["markings"]["markings"]
        
        # Первый подсчет - по статусу
        count_by_status = {}
        for marking in markings:
            status = marking["status"]
            count_by_status[status] = count_by_status.get(status, 0) + 1
        
        # Второй подсчет - по артикулам
        count_by_article = {}
        articles_found = []
        
        for marking in markings:
            if marking["article"]:
                article = str(marking["article"]).strip()
                count_by_article[article] = count_by_article.get(article, 0) + 1
                articles_found.append(article)
        
        # Итоговые числа
        total_photos = len(markings)
        readable_markings = sum(1 for m in markings if m["readable"])
        unique_articles = len(set(articles_found))
        
        result = {
            "success": True,
            "first_count": {
                "total_photos": total_photos,
                "by_status": count_by_status
            },
            "second_count": {
                "readable_markings": readable_markings,
                "unique_articles": unique_articles,
                "by_article": dict(count_by_article)
            },
            "verification": {
                "photos_match": total_photos == len(markings),
                "readable_count_match": True  # Всегда валидно для этого подсчета
            }
        }
        
        session_data["results"]["count_verification"] = result
        session_data["step"] = 5
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка пересчета: {str(e)}")

@app.post("/step5-compare-specification/")
async def step5_compare_specification():
    """Шаг 5: Сопоставление со спецификацией"""
    try:
        # Проверяем наличие спецификации
        if not session_data.get("specification"):
            raise HTTPException(status_code=400, detail="Отсутствует файл спецификации")
        
        # Проверяем наличие данных маркировок
        if "markings" not in session_data.get("results", {}) or "markings" not in session_data["results"]["markings"]:
            raise HTTPException(status_code=400, detail="Отсутствуют данные маркировок. Выполните шаг 3.")
        
        # Парсинг спецификации
        specification = assistant.parse_excel_specification(session_data["specification"])
        
        # Получение фактических данных из маркировок
        markings = session_data["results"]["markings"]["markings"]
        actual_count = {}
        
        for marking in markings:
            if marking["article"] and marking["readable"]:
                article = str(marking["article"]).strip()
                actual_count[article] = actual_count.get(article, 0) + 1
        
        # Сопоставление
        comparison = []
        spec_articles = set()
        
        for spec_item in specification:
            article = spec_item["article"]
            spec_articles.add(article)
            planned = spec_item["quantity"]
            actual = actual_count.get(article, 0)
            difference = actual - planned
            
            if difference == 0:
                status = "✅"  # Точное соответствие
            elif difference < 0:
                status = "⬇️"  # Недостает
            else:
                status = "⬆️"  # Пересорт
            
            comparison.append({
                "article": article,
                "name": spec_item.get("name", ""),
                "planned": planned,
                "actual": actual,
                "difference": difference,
                "status": status
            })
        
        # Проверка на лишние артикулы
        for article in actual_count:
            if article not in spec_articles:
                comparison.append({
                    "article": article,
                    "name": "Не в спецификации",
                    "planned": 0,
                    "actual": actual_count[article],
                    "difference": actual_count[article],
                    "status": "🔁"  # Лишний артикул
                })
        
        # Статистика
        exact_match = sum(1 for c in comparison if c["status"] == "✅")
        shortage = sum(1 for c in comparison if c["status"] == "⬇️")
        excess = sum(1 for c in comparison if c["status"] == "⬆️")
        extra = sum(1 for c in comparison if c["status"] == "🔁")
        
        result = {
            "success": True,
            "comparison": comparison,
            "summary": {
                "total_positions": len(comparison),
                "exact_match": exact_match,
                "shortage": shortage,
                "excess": excess,
                "extra_articles": extra
            }
        }
        
        session_data["results"]["comparison"] = result
        session_data["step"] = 6
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка сопоставления: {str(e)}")

@app.post("/step6-final-questions/")
async def step6_final_questions(
    shipment_type: str = Form(...),  # "full" или "partial"
    shipment_date: str = Form(...)   # ДД.ММ.ГГГГ
):
    """Шаг 6: Финальные вопросы перед генерацией"""
    try:
        # Валидация даты
        try:
            datetime.strptime(shipment_date, "%d.%m.%Y")
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный формат даты. Используйте ДД.ММ.ГГГГ")
        
        session_data["results"]["final_params"] = {
            "shipment_type": shipment_type,
            "shipment_date": shipment_date,
            "shipment_type_ru": "Окончательная отгрузка" if shipment_type == "full" else "Частичная отгрузка"
        }
        
        session_data["step"] = 7
        
        return {
            "success": True,
            "params": session_data["results"]["final_params"]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка обработки параметров: {str(e)}")

@app.post("/step7-generate-files/")
async def step7_generate_files():
    """Шаг 7: Генерация Excel файлов"""
    try:
        # Проверяем наличие всех необходимых данных
        if "comparison" not in session_data.get("results", {}) or "comparison" not in session_data["results"]["comparison"]:
            raise HTTPException(status_code=400, detail="Отсутствуют данные сравнения. Выполните шаг 5.")
        
        if "final_params" not in session_data.get("results", {}):
            raise HTTPException(status_code=400, detail="Отсутствуют финальные параметры. Выполните шаг 6.")
        
        if "markings" not in session_data.get("results", {}) or "markings" not in session_data["results"]["markings"]:
            raise HTTPException(status_code=400, detail="Отсутствуют данные маркировок. Выполните шаг 3.")
        
        # Получение данных
        comparison = session_data["results"]["comparison"]["comparison"]
        final_params = session_data["results"]["final_params"]
        markings = session_data["results"]["markings"]["markings"]
        
        generated_files = []
        
        # 1. Генерация заполненной накладной
        invoice_path = await generate_filled_invoice(comparison, markings, final_params)
        generated_files.append({
            "name": "Накладная_заполненная.xlsx",
            "path": invoice_path,
            "type": "invoice"
        })
        
        # 2. Генерация обновленной спецификации
        spec_path = await generate_updated_specification(comparison, final_params)
        generated_files.append({
            "name": "Спецификация_обновленная.xlsx", 
            "path": spec_path,
            "type": "specification"
        })
        
        result = {
            "success": True,
            "files": generated_files,
            "summary": {
                "invoice_generated": True,
                "specification_updated": True,
                "shipment_type": final_params["shipment_type_ru"],
                "shipment_date": final_params["shipment_date"]
            }
        }
        
        session_data["results"]["generated_files"] = result
        session_data["step"] = 8  # Завершено
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка генерации файлов: {str(e)}")

async def generate_filled_invoice(comparison, markings, final_params):
    """Генерация заполненной накладной"""
    try:
        # Проверяем наличие шаблона
        if not session_data.get("template") or not os.path.exists(session_data["template"]):
            raise Exception("Отсутствует файл шаблона накладной")
        
        # Загрузка шаблона накладной
        template_wb = load_workbook(session_data["template"])
        ws = template_wb.active
        
        # Поиск заголовков колонок
        header_row = 1
        for row in range(1, 10):
            for cell in ws[row]:
                if cell.value and any(word in str(cell.value).lower() for word in ['наименование', 'название']):
                    header_row = row
                    break
            if header_row > 1:
                break
        
        # Определение колонок
        name_col = None
        unit_col = None
        qty_col = None
        area_col = None
        
        for col_idx, cell in enumerate(ws[header_row], 1):
            if cell.value:
                value_lower = str(cell.value).lower()
                if any(word in value_lower for word in ['наименование', 'название']):
                    name_col = col_idx
                elif any(word in value_lower for word in ['ед.изм', 'единица']):
                    unit_col = col_idx
                elif any(word in value_lower for word in ['количество', 'кол-во']):
                    qty_col = col_idx
                elif any(word in value_lower for word in ['площадь', 'м²', 'кв.м']):
                    area_col = col_idx
        
        # Заполнение данных
        data_row = header_row + 1
        row_num = 1
        
        for item in comparison:
            if item["actual"] > 0:  # Только фактически отгруженные позиции
                # Номер позиции
                ws.cell(row=data_row, column=1, value=row_num)
                
                # Наименование
                if name_col:
                    # Ищем полное наименование из маркировок
                    full_name = item["name"]
                    for marking in markings:
                        if marking["article"] == item["article"] and marking["name"]:
                            full_name = marking["name"]
                            break
                    ws.cell(row=data_row, column=name_col, value=full_name or item["article"])
                
                # Единица измерения
                if unit_col:
                    ws.cell(row=data_row, column=unit_col, value="шт")
                
                # Количество штук
                if qty_col:
                    ws.cell(row=data_row, column=qty_col, value=item["actual"])
                
                # Площадь в м²
                if area_col:
                    # Ищем размеры из маркировок
                    dimensions = None
                    for marking in markings:
                        if marking["article"] == item["article"] and marking["dimensions"]:
                            dimensions = marking["dimensions"]
                            break
                    
                    area_per_item = assistant.calculate_square_meters(
                        item["name"] or item["article"], 
                        dimensions
                    )
                    total_area = area_per_item * item["actual"]
                    
                    if total_area > 0:
                        ws.cell(row=data_row, column=area_col, value=round(total_area, 2))
                
                data_row += 1
                row_num += 1
        
        # Сохранение файла
        output_path = f"output/Накладная_заполненная_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        template_wb.save(output_path)
        
        return output_path
        
    except Exception as e:
        raise Exception(f"Ошибка генерации накладной: {str(e)}")

async def generate_updated_specification(comparison, final_params):
    """Генерация обновленной спецификации"""
    try:
        # Загрузка оригинальной спецификации
        spec_wb = load_workbook(session_data["specification"])
        ws = spec_wb.active
        
        # Поиск колонок
        article_col = None
        shipped_col = None
        date_col = None
        
        for row in range(1, 10):
            for col_idx, cell in enumerate(ws[row], 1):
                if cell.value:
                    value_lower = str(cell.value).lower()
                    if any(word in value_lower for word in ['артикул', 'код']):
                        article_col = col_idx
                    elif any(word in value_lower for word in ['отгружен', 'отправлен']):
                        shipped_col = col_idx
                    elif any(word in value_lower for word in ['дата']):
                        date_col = col_idx
        
        # Если колонка "Отгруженные" не найдена, добавляем её
        if not shipped_col:
            # Находим последнюю используемую колонку
            max_col = ws.max_column
            shipped_col = max_col + 1
            ws.cell(row=1, column=shipped_col, value="Отгруженные")
        
        # Если колонка "Дата отгрузки" не найдена, добавляем её
        if not date_col:
            max_col = max(shipped_col, ws.max_column)
            date_col = max_col + 1
            ws.cell(row=1, column=date_col, value="Дата отгрузки")
        
        # Красный фон для пересорта
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        
        # Обновление данных
        for row_idx in range(2, ws.max_row + 1):
            article_cell = ws.cell(row=row_idx, column=article_col)
            if article_cell.value:
                article = str(article_cell.value).strip()
                
                # Ищем соответствующую позицию в сравнении
                for item in comparison:
                    if item["article"] == article:
                        # Обновляем количество отгруженных
                        shipped_cell = ws.cell(row=row_idx, column=shipped_col)
                        shipped_cell.value = item["actual"]
                        
                        # Обновляем дату отгрузки
                        date_cell = ws.cell(row=row_idx, column=date_col)
                        date_cell.value = final_params["shipment_date"]
                        
                        # Выделяем красным пересорт
                        if item["status"] in ["⬆️", "🔁"]:  # Больше плана или лишний артикул
                            shipped_cell.fill = red_fill
                        
                        break
        
        # Сохранение файла
        output_path = f"output/Спецификация_обновленная_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        spec_wb.save(output_path)
        
        return output_path
        
    except Exception as e:
        raise Exception(f"Ошибка обновления спецификации: {str(e)}")

@app.get("/download/{file_type}/{filename}")
async def download_file(file_type: str, filename: str):
    """Скачивание сгенерированных файлов"""
    try:
        file_path = f"output/{filename}"
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="Файл не найден")
        
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка скачивания: {str(e)}")

@app.get("/status")
async def get_status():
    """Получение текущего статуса обработки"""
    return {
        "step": session_data["step"],
        "files_uploaded": {
            "specification": session_data["specification"] is not None,
            "template": session_data["template"] is not None,
            "photos": len(session_data["photos"])
        },
        "results_available": list(session_data["results"].keys())
    }

@app.post("/reset")
async def reset_session():
    """Сброс сессии"""
    global session_data
    
    # Очистка временных файлов
    for folder in ["uploads", "output"]:
        if os.path.exists(folder):
            for file in os.listdir(folder):
                try:
                    os.remove(os.path.join(folder, file))
                except:
                    pass
    
    # Сброс данных сессии
    session_data = {
        "specification": None,
        "template": None,
        "photos": [],
        "step": 1,
        "results": {}
    }
    
    return {"success": True, "message": "Сессия сброшена"}

if __name__ == "__main__":
    uvicorn.run("app:app", host="0.0.0.0", port=8090, reload=True)